from __future__ import annotations

import asyncio
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import AsyncSessionLocal
from app.models import Customer, Product, Store

DESKTOP_DIR = Path.home() / "Desktop"
PRODUCT_XLSX = DESKTOP_DIR / "商品信息.xlsx"
CUSTOMER_XLSX = DESKTOP_DIR / "客户信息表.xlsx"
PRODUCT_LIMIT = 50
CUSTOMER_LIMIT = 50


@dataclass
class ImportSummary:
    imported_products: int = 0
    imported_customers: int = 0


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_decimal(value: Any, default: Decimal = Decimal("0.00")) -> Decimal:
    text = normalize_text(value).replace(",", "")
    if not text:
        return default
    try:
        return Decimal(text)
    except InvalidOperation:
        return default


def parse_bool(value: Any) -> bool:
    return normalize_text(value) in {"是", "true", "True", "1", "Y", "YES"}


def parse_int(value: Any) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_birth_to_age(value: Any) -> int | None:
    if not value:
        return None
    if isinstance(value, datetime):
        born = value.date()
    elif isinstance(value, date):
        born = value
    else:
        text = normalize_text(value)
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                born = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        else:
            return None

    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


def iter_product_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(PRODUCT_XLSX, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))]
    rows: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        payload = {headers[index]: row[index] for index in range(len(headers))}
        rows.append(payload)

    workbook.close()
    return rows


def iter_customer_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(CUSTOMER_XLSX, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    rows: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        payload = {headers[index]: row[index] for index in range(len(headers))}
        rows.append(payload)

    workbook.close()
    return rows


def build_product_category(row: dict[str, Any]) -> str:
    category_parts = [
        normalize_text(row.get("*一级类别")),
        normalize_text(row.get("二级类别")),
        normalize_text(row.get("三级类别")),
    ]
    category_parts = [part for part in category_parts if part]
    if category_parts:
        return " / ".join(category_parts)

    for key in ("*商品类型", "商品类型", "规格型号"):
        value = normalize_text(row.get(key))
        if value:
            return value

    return "未分类"


async def import_products() -> int:
    rows = iter_product_rows()

    async with AsyncSessionLocal() as session:
        existing_skus = {
            sku
            for sku in (
                await session.execute(select(Product.sku))
            ).scalars()
        }

        new_products: list[Product] = []
        seen_skus = set(existing_skus)

        for row in rows:
            sku = normalize_text(row.get("*商品编号"))
            name = normalize_text(row.get("*商品名称"))
            retail_price = parse_decimal(row.get("*零售价"))

            if not sku or not name or retail_price <= 0:
                continue
            if sku in seen_skus:
                continue

            cost_price = parse_decimal(row.get("成本价"))
            if cost_price <= 0:
                cost_price = (retail_price * Decimal("0.6")).quantize(Decimal("0.01"))

            product = Product(
                sku=sku,
                name=name,
                category=build_product_category(row),
                retail_price=retail_price,
                cost_price=cost_price,
                brand=normalize_text(row.get("品牌")) or None,
                manufacturer=normalize_text(row.get("生产厂家")) or normalize_text(row.get("供应商")) or None,
                registration_no=normalize_text(row.get("注册证号")) or None,
                has_sn_tracking=parse_bool(row.get("*启用机编")),
            )
            new_products.append(product)
            seen_skus.add(sku)

            if len(new_products) >= PRODUCT_LIMIT:
                break

        session.add_all(new_products)
        await session.commit()
        return len(new_products)


async def import_customers() -> int:
    rows = iter_customer_rows()

    async with AsyncSessionLocal() as session:
        store_rows = (await session.execute(select(Store))).scalars().all()
        if not store_rows:
            raise RuntimeError("No stores found. Seed stores first.")

        default_store = store_rows[0]
        store_map = {store.name: store for store in store_rows}

        existing_customer_keys = {
            (name, phone)
            for name, phone in (
                await session.execute(select(Customer.name, Customer.phone))
            ).all()
        }

        new_customers: list[Customer] = []
        seen_customer_keys = set(existing_customer_keys)

        for row in rows:
            name = normalize_text(row.get("客户姓名"))
            phone = normalize_text(row.get("手机号码")) or normalize_text(row.get("电话号码"))

            if not name or not phone:
                continue

            customer_key = (name, phone)
            if customer_key in seen_customer_keys:
                continue

            store_name = normalize_text(row.get("所属门店"))
            mapped_store = store_map.get(store_name, default_store)
            age = parse_int(row.get("年龄")) or parse_birth_to_age(row.get("出生日期"))

            customer = Customer(
                name=name,
                phone=phone,
                age=age,
                gender=normalize_text(row.get("性别")) or None,
                hearing_loss_type=None,
                primary_store_id=mapped_store.id,
                ocr_raw_data={
                    "source": "desktop_excel_import",
                    "customer_no": normalize_text(row.get("客户编号")) or None,
                    "province": normalize_text(row.get("省份")) or None,
                    "city": normalize_text(row.get("城市")) or None,
                    "district": normalize_text(row.get("区/县")) or None,
                    "address": normalize_text(row.get("详细地址")) or None,
                    "store_name": store_name or mapped_store.name,
                    "owner": normalize_text(row.get("所属人员")) or None,
                    "notes": normalize_text(row.get("备注信息")) or None,
                },
            )
            new_customers.append(customer)
            seen_customer_keys.add(customer_key)

            if len(new_customers) >= CUSTOMER_LIMIT:
                break

        session.add_all(new_customers)
        await session.commit()
        return len(new_customers)


async def main() -> None:
    summary = ImportSummary()
    summary.imported_products = await import_products()
    summary.imported_customers = await import_customers()

    print(f"Imported products: {summary.imported_products}")
    print(f"Imported customers: {summary.imported_customers}")


if __name__ == "__main__":
    asyncio.run(main())
